#!/usr/bin/python

import sys
import os.path
import argparse
import openpyxl
import re

## Function: A closure for file extension checking

def ext_check(expected_ext, openner):
	def extension(filename):
		if not filename.lower().endswith(expected_ext):
			raise ValueError()
		return openner(filename)
	return extension


parser = argparse.ArgumentParser(description='Print out a selected worksheet from .xlsx workbook')

parser.add_argument("workbook", type=ext_check('.xlsx', argparse.FileType('r')))

args = parser.parse_args()

workBookPath = args.workbook.name

qcWorkbook = openpyxl.load_workbook(workBookPath)

print(qcWorkbook.sheetnames)

sheet =qcWorkbook['AZ']

dataPresent = True
dataCount = 0
offset=0

wsRow=1
wsCol=1
qR1=""
qR2=""
coverage=""
genomeLen=""
insert=""
hqSNP=""
name=""

print("<html><body>")

while(dataPresent is True):
	if( (sheet.cell(row=wsRow, column=1).value is None) and (sheet.cell(row=wsRow, column=2).value is None) ):
		#print(sheet.cell(row=wsRow, column=1).value, end=" ")
		dataCount = dataCount + 1
	else:
		dataCount = 0
		if(sheet.cell(row=wsRow, column=1).value == 'R1'):
			name = str(sheet.cell(row=wsRow, column=2).value)
			print("<table style=\"font-size:13px; text-align:center\">")
			qR1 = "<tr><td><b>" + str(sheet.cell(row=wsRow, column=3).value) + "</b> (min 30)</td></tr>"
			print(qR1)
		elif(sheet.cell(row=wsRow, column=1).value == 'R2'):
			name = name + "\t" + str(sheet.cell(row=wsRow, column=2).value)
			qR2 = "<tr><td><b>" + str(sheet.cell(row=wsRow, column=3).value) + "</b> (min 30)</td></tr>"
			print(qR2)
		elif(re.search('depth', str(sheet.cell(row=wsRow-1 , column=2).value), flags=re.IGNORECASE) ):
			for ii in range(1, 12):
				#print(sheet.cell(row=wsRow, column=ii).value, end=" ")
				if(re.search('depth', str(sheet.cell(row=wsRow-1 , column=ii).value), flags=re.IGNORECASE) ):
					floatCover = round(float(sheet.cell(row=wsRow, column=ii).value), 0)
					coverage = "<tr><td><b>" + str(int(floatCover)) + "x</b> (min 20x)</td></tr>"
					print(coverage)
				if(re.search(r'MedianInsert', str(sheet.cell(row=wsRow-1 , column=ii).value), flags=re.IGNORECASE)):
					floatInsert = round(float(sheet.cell(row=wsRow, column=ii).value), 0)
					insert = "<tr><td><b>" + str(int(floatInsert)) + "</b> median (median > 300bp)</td></tr>"
					print(insert)
				if(re.search(r'total', str(sheet.cell(row=wsRow-1, column=ii).value), flags=re.IGNORECASE)):
					floatGenomeLen = round(float(sheet.cell(row=wsRow, column=ii).value)/1000000, 1)
					genomeLen = "<tr><td><b>" + str(floatGenomeLen) + "MB</b> (1.7MB, 5% error margin)</td></tr>"
		elif(re.search('D5480', str(sheet.cell(row=wsRow , column=1).value)) or re.search('D5480', str(sheet.cell(row=wsRow ,column=2).value))):
			hqSNP="<tr><td><b>" + str(sheet.cell(row=wsRow , column=3).value) + "</b> (<1/ MB)</td></tr>"	
			print(hqSNP)
			print("<tr><td>NA</td></tr>")
			print(genomeLen)
			print("</table>")
			print()
			print(name)
	wsRow = wsRow + 1
	if(dataCount > 10):
		dataPresent = False

print("</body></html>")

qcWorkbook.save(workBookPath)
qcWorkbook.close()



	
